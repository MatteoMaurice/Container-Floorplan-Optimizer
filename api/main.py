from __future__ import annotations
import csv
import io
import json
from typing import Any, Dict, List, Optional

from fastapi import FastAPI, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field

from core.optimize import optimize_layout
from openpyxl import load_workbook

app = FastAPI(title="Container Floorplan Optimizer")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def load_containers() -> Dict[str, Any]:
    with open("api/containers.json", "r", encoding="utf-8") as f:
        return json.load(f)

CONTAINERS = load_containers()

class OptimizeRequest(BaseModel):
    container_code: str = Field(..., description="e.g. 20GP, 40GP, 40HC")
    margin_mm: int = 0
    compare_all: bool = False
    override_L_mm: Optional[int] = None
    override_W_mm: Optional[int] = None
    rows: List[Dict[str, Any]]  # parsed CSV rows

@app.get("/containers")
def get_containers():
    return CONTAINERS

def parse_csv_bytes(data: bytes) -> List[Dict[str, Any]]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for r in reader:
        # normalize keys
        def g(*keys, default=None):
            for k in keys:
                if k in r and r[k] not in (None, ""):
                    return r[k]
            return default

        sku = g("sku", "name", "id", default="ITEM")
        qty = int(g("qty", "quantity", default="1"))
        length = int(g("length_mm", "length", "L", "l"))
        width = int(g("width_mm", "width", "W", "w"))
        rot = g("rotatable", "rotate", default="true")
        rows.append({
            "sku": sku,
            "qty": qty,
            "length_mm": length,
            "width_mm": width,
            "rotatable": rot,
        })
    return rows
def parse_excel_bytes(data: bytes) -> List[Dict[str, Any]]:
    wb = load_workbook(filename=io.BytesIO(data), data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel file appears empty.")

    header = [str(c or "").strip() for c in rows[0]]

    def norm(s: str) -> str:
        return (s or "").strip().lower().replace(" ", "").replace("-", "").replace("_", "")

    idx_map = {norm(h): i for i, h in enumerate(header)}

    def get_cell(r, *names):
        for n in names:
            i = idx_map.get(norm(n))
            if i is not None and i < len(r):
                v = r[i]
                if v is not None and str(v).strip() != "":
                    return v
        return None

    out: List[Dict[str, Any]] = []
    for line_no, r in enumerate(rows[1:], start=2):
        if r is None or all((c is None or str(c).strip() == "") for c in r):
            continue

        sku = get_cell(r, "sku", "name", "id") or "ITEM"
        qty = get_cell(r, "qty", "quantity", "qte", "count") or 1
        length = get_cell(r, "length_mm", "length", "L", "l", "longueur", "lenght")
        width  = get_cell(r, "width_mm", "width", "W", "w", "largeur", "wistd", "with")

        if length is None or width is None:
            raise ValueError(f"Excel parse error line {line_no}: missing length/width columns.")

        out.append({
            "sku": str(sku).strip(),
            "qty": int(float(qty)),
            "length_mm": int(float(length)),
            "width_mm": int(float(width)),
            "rotatable": str(get_cell(r, "rotatable", "rotate", "rotation") or "true").strip(),
        })

    if not out:
        raise ValueError("Excel file has headers but no data rows.")
    return out
@app.post("/optimize")
def optimize(req: OptimizeRequest):
    def resolve_dims(code: str):
        c = CONTAINERS.get(code)
        if not c:
            raise ValueError(f"Unknown container_code: {code}")
        L = int(req.override_L_mm or c["L_mm"])
        W = int(req.override_W_mm or c["W_mm"])
        return L, W, c

    if req.compare_all:
        results = {}
        for code in CONTAINERS.keys():
            L, W, c = resolve_dims(code)
            results[code] = {
                "meta": c,
                "result": optimize_layout(L, W, req.rows, margin=req.margin_mm)
            }
        # pick best: max placed, then max utilization
        best_code = max(
            results.keys(),
            key=lambda k: (
                results[k]["result"]["summary"]["placed"],
                results[k]["result"]["summary"]["utilization_area"]
            )
        )
        return {"mode": "compare_all", "best_container": best_code, "results": results}

    L, W, c = resolve_dims(req.container_code)
    res = optimize_layout(L, W, req.rows, margin=req.margin_mm)
    return {"mode": "single", "container_code": req.container_code, "meta": c, "result": res}

@app.post("/optimize_csv")
async def optimize_csv(
    container_code: str,
    margin_mm: int = 0,
    compare_all: bool = False,
    file: UploadFile = File(...)
):
    data = await file.read()
    name = (file.filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        rows = parse_excel_bytes(data)
    else:
        rows = parse_csv_bytes(data)
    req = OptimizeRequest(
        container_code=container_code,
        margin_mm=margin_mm,
        compare_all=compare_all,
        rows=rows
    )
    return optimize(req)
